import requests
from bs4 import BeautifulSoup as bs
import openpyxl as opx
from openpyxl.styles import PatternFill
# import the necessary 3rd party libraries
redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')
# load the already saved workbook
# wb = opx.load_workbook('Google Code-in Preparation.xlsx')
# or make a new one
wb = opx.Workbook()

# for downloading every year from 2015 to 2018
for year in range(2015, 2019):
    # create a new sheet with the year as title
    sheet = wb.create_sheet(title=str(year))
    # name the first cell as organisation
    sheet.cell(row=1, column=1).value = 'Organisation'
    # freeze the row A and column 1
    sheet.freeze_panes = 'B2'
    # all suburls begin from this
    baseurl = 'https://codein.withgoogle.com'
    # load the respective archive page and make the soup and find the org link
    res = requests.get('https://codein.withgoogle.com/archive/{}/'.format(str(year)))
    soup = bs(res.text, 'lxml')
    org = soup.select('div > h3 > a')
    # as first row is for programmin lang names so start from 2
    orgindex = 2
    # to save time, so instead of searching for row make column and save its id
    langdict = {}
    langcount = {}
    #for every org in that list
    for o in org:
        print(o.text)
        sheet['A' + str(orgindex)] = o.text
        #get the org page and make soup and find the programmin lang list
        res = requests.get(baseurl + o['href'])
        soup = bs(res.text, 'lxml')
        prolang = soup.select('.org__tag')

        for pl in prolang:
            print(pl.text)
            try:
                #if the lang is already present find for its responding column and fill it
                sheet.cell(row=orgindex, column=langdict[pl.text]).value = pl.text
                sheet.cell(row=orgindex, column=langdict[pl.text]).fill = redFill
                langcount[pl.text] += 1

            except Exception as e:
                print(e)
                #if the lang is not found then it adds the lang in the dict and runs again
                langdict[pl.text] = len(langdict.items()) + 2
                langcount[pl.text] = 1

                sheet.cell(row=1, column=langdict[pl.text]).value = pl.text

                sheet.cell(row=orgindex, column=langdict[pl.text]).value = pl.text
                sheet.cell(row=orgindex, column=langdict[pl.text]).fill = redFill

        orgindex += 1
    #after this year's org list finishes add the total count row
    sheet['A'+str(orgindex)] = 'Total Count'
    for lang in langdict.keys():
        sheet.cell(row=orgindex, column=langdict[lang]).value = langcount[lang]
    print(langcount)
    print(langdict)

#most important part save the workbook
wb.save('Google Code-in Preparation.xlsx')
