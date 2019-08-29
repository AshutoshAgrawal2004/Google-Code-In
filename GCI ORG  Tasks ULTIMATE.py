import time

# for performance tests record time and import other modules
starttime = time.time()

import requests as req
from bs4 import BeautifulSoup as bs
import openpyxl as opx
import logging
import re
from collections import Counter

# setup logger
logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s -%(levelname)s - %(message)s')

# set up the workbook
wb = opx.Workbook()
sheet = wb['Sheet']
sheet['A1'] = 'Task'
sheet['B1'] = 'Link'
sheet['C1'] = 'Solver'
sheet['D1'] = 'Organization'

sheet.freeze_panes = 'D2'

# basically its the current row number
taskindex = 2
tasktakers = []
orgtakers = []
tasktypedict = {}

# parse and find the url from input
baseurl = 'https://codein.withgoogle.com'
orgurl = 'https://codein.withgoogle.com/archive/2018/task/'
# initially get the total pages
res = req.get(orgurl)
soup = bs(res.text, 'lxml')
totalpagepattern = re.compile(r'Page\s(\d+)\sof\s(\d+)')
paginator = soup.select_one('.paginator__pages').text
current_page, total_page = map(int, re.search(totalpagepattern, paginator).groups())


def getTasks(current_page):
    print('Page {} of {}'.format(current_page, total_page))
    global taskindex
    # get the task page of the currentpagenumber and get all the links
    res = req.get(orgurl + '?page=' + str(current_page))
    soup = bs(res.text, 'lxml')
    tasklinks = soup.select('.md-raised')
    for task in tasklinks:
        # go to each task page and get the respective info
        res = req.get(baseurl + task['href'])
        soup = bs(res.text, 'lxml')
        taskname = soup.select_one('.task-definition__name')
        orgname = soup.select_one('.task-definition__organization')
        doneby = soup.select_one('.task-definition__students-subheader+ div').text.encode("utf-8")
        # as number of solvers can be many we convert into an array for counting purpose
        orgtakers.append(orgname.text)
        l = str(doneby).split(',')
        c = []
        print(l)
        for participant in l:
            c.append(participant.strip(" 'b"))
            tasktakers.append(participant.strip(" 'b"))
        # now write the basic info on the spreadsheet
        try:
            sheet['A' + str(taskindex)] = taskname.text
            sheet['B' + str(taskindex)] = baseurl + task['href']
            sheet['C' + str(taskindex)] = ", ".join(c)
            sheet['D' + str(taskindex)] = orgname.text

            logging.info(taskname.text)
        except Exception as e:
            logging.error(e)
            print(taskname)

        # now get the tasktype and loop through it
        tasktype = soup.select('.task-category__name')

        for tt in tasktype:
            logging.info(tt.text)
            try:
                # if its already present fill its value
                sheet.cell(row=taskindex, column=tasktypedict[tt.text]).value = tt.text
            except Exception as e:
                # else add in dict then fill the value
                logging.error(e)
                tasktypedict[tt.text] = len(tasktypedict.items()) + 5
                sheet.cell(row=1, column=tasktypedict[tt.text]).value = tt.text

                sheet.cell(row=taskindex, column=tasktypedict[tt.text]).value = tt.text
        # increment the task index basically its the row number
        taskindex += 1
        print(taskindex)
    # after the page finishes move to next page if its available
    if current_page < total_page:
        getTasks(current_page + 1)


# call the function initially
getTasks(current_page)

# now use the counter to count number of tasks completed by each
tasktakersStats = Counter(tasktakers)
orgtakersStats = Counter(orgtakers)
# finally save the workbook and print all the collected info
wb.save('GCI All Tasks.xlsx')
endtime = time.time()
print(tasktypedict)
print(tasktakers)
print(tasktakersStats)
print(orgtakers)
print(orgtakersStats)
print('Processing took {}'.format(endtime - starttime))
